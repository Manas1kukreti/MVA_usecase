"""XLSX file loader — enforces single-sheet rule and validates content."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.config import Settings
from app.core.exceptions import FileValidationError, DatasetLimitError
from app.core.logging import get_logger

logger = get_logger(__name__)


class XLSXLoader:
    """Load and validate XLSX files into a pandas DataFrame."""

    def __init__(self, settings: Settings):
        self._max_rows = settings.max_dataset_rows
        self._max_columns = settings.max_dataset_columns

    def _get_non_empty_sheets(self, file_path: Path) -> list[str]:
        """
        Identify non-empty tabular sheets in the workbook.

        A sheet is considered non-empty if it has at least one row of data
        beyond the header.
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        non_empty = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_count = 0
            for row in ws.iter_rows(max_row=3):  # Check first few rows
                if any(cell.value is not None for cell in row):
                    row_count += 1
            if row_count >= 2:  # At least header + one data row
                non_empty.append(sheet_name)

        wb.close()
        return non_empty

    def load(self, file_path: Path) -> pd.DataFrame:
        """
        Load an XLSX file into a DataFrame.

        Validates:
        - File is readable
        - Workbook contains exactly one non-empty tabular sheet
        - Sheet has a header row
        - Does not exceed row/column limits

        Returns the loaded DataFrame.
        """
        if not file_path.exists():
            raise FileValidationError(
                code="FILE_NOT_FOUND",
                message="The file could not be found for processing.",
                details={"path": str(file_path)},
            )

        try:
            non_empty_sheets = self._get_non_empty_sheets(file_path)
        except Exception as e:
            raise FileValidationError(
                code="XLSX_READ_ERROR",
                message=f"Failed to read XLSX file: {str(e)}",
                details={"error": str(e)},
            )

        if len(non_empty_sheets) == 0:
            raise FileValidationError(
                code="EMPTY_WORKBOOK",
                message="The workbook contains no non-empty tabular sheets.",
                details={"sheet_names": []},
            )

        if len(non_empty_sheets) > 1:
            raise FileValidationError(
                code="MULTIPLE_XLSX_SHEETS",
                message="The workbook contains more than one non-empty tabular sheet.",
                details={"sheet_names": non_empty_sheets},
            )

        sheet_name = non_empty_sheets[0]
        logger.info("xlsx_sheet_selected", sheet_name=sheet_name)

        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                dtype=str,  # Load all as string initially for type refinement
                na_values=["", "NA", "N/A", "null", "NULL", "None", "none", "nan", "NaN"],
            )
        except Exception as e:
            raise FileValidationError(
                code="XLSX_PARSE_ERROR",
                message=f"Failed to parse XLSX sheet '{sheet_name}': {str(e)}",
                details={"sheet_name": sheet_name, "error": str(e)},
            )

        self._validate_dataframe(df)
        return df

    def _validate_dataframe(self, df: pd.DataFrame) -> None:
        """Validate the loaded dataframe against configured limits."""
        if df.empty or len(df.columns) == 0:
            raise FileValidationError(
                code="EMPTY_DATASET",
                message="The sheet contains no data or no columns.",
                details={"rows": len(df), "columns": len(df.columns)},
            )

        if len(df) == 0:
            raise FileValidationError(
                code="NO_DATA_ROWS",
                message="The sheet contains a header but no data rows.",
                details={"columns": len(df.columns)},
            )

        if len(df) > self._max_rows:
            raise DatasetLimitError(
                code="DATASET_ROW_LIMIT_EXCEEDED",
                message=(
                    f"The dataset contains {len(df):,} rows. "
                    f"The configured maximum is {self._max_rows:,} rows."
                ),
                details={"actual_rows": len(df), "maximum_rows": self._max_rows},
            )

        if len(df.columns) > self._max_columns:
            raise DatasetLimitError(
                code="DATASET_COLUMN_LIMIT_EXCEEDED",
                message=(
                    f"The dataset contains {len(df.columns)} columns. "
                    f"The configured maximum is {self._max_columns} columns."
                ),
                details={"actual_columns": len(df.columns), "maximum_columns": self._max_columns},
            )
