"""Integration tests for the API endpoints."""

import pytest
import json
from fastapi.testclient import TestClient

from app.main import create_app


@pytest.fixture
def client():
    app = create_app()
    with TestClient(app) as c:
        yield c


@pytest.fixture
def csv_content() -> bytes:
    lines = ["txn_id,amount,status,auth_date,country"]
    for i in range(1, 31):
        status = "approved" if i % 3 != 0 else "declined"
        lines.append(f"T{i:04d},{i * 10.5},{status},2024-01-{(i % 28) + 1:02d},US")
    return "\n".join(lines).encode("utf-8")


class TestHealthEndpoint:
    def test_health(self, client):
        resp = client.get("/api/v1/health")
        assert resp.status_code == 200
        data = resp.json()
        assert data["pipeline_version"] == "1.0.0"


class TestProfileRunEndpoints:
    def test_create_profile_run_csv(self, client, csv_content):
        resp = client.post(
            "/api/v1/profile-runs",
            data={
                "primary_domain": "Payments",
                "schema_metadata": json.dumps({"columns": [
                    {"column_name": "amount", "description": "Payment amount", "mandatory": True, "expected_unique": False},
                    {"column_name": "txn_id", "description": "Transaction ID", "mandatory": False, "expected_unique": True},
                ]}),
            },
            files={"file": ("payments.csv", csv_content, "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "completed"
        run_id = data["run_id"]

        # Get summary
        resp2 = client.get(f"/api/v1/profile-runs/{run_id}")
        assert resp2.status_code == 200
        summary = resp2.json()
        assert summary["primary_domain"] == "Payments"

        # Get full result
        resp3 = client.get(f"/api/v1/profile-runs/{run_id}/result")
        assert resp3.status_code == 200
        result = resp3.json()
        assert len(result["column_profiles"]) == 5
        assert len(result["quality_assessments"]) >= 9
        assert len(result["readiness_assessments"]) == 4

        # Get columns
        resp4 = client.get(f"/api/v1/profile-runs/{run_id}/columns")
        assert resp4.status_code == 200
        assert len(resp4.json()["columns"]) == 5

        # Get quality
        resp5 = client.get(f"/api/v1/profile-runs/{run_id}/quality")
        assert resp5.status_code == 200
        assert "overall" in resp5.json()

        # Get readiness
        resp6 = client.get(f"/api/v1/profile-runs/{run_id}/readiness")
        assert resp6.status_code == 200
        assert len(resp6.json()["assessments"]) == 4

        # Get charts
        resp7 = client.get(f"/api/v1/profile-runs/{run_id}/charts")
        assert resp7.status_code == 200

        # Get hierarchy
        resp8 = client.get(f"/api/v1/profile-runs/{run_id}/hierarchy")
        assert resp8.status_code == 200

    def test_unsupported_domain(self, client, csv_content):
        resp = client.post(
            "/api/v1/profile-runs",
            data={"primary_domain": "Insurance"},
            files={"file": ("data.csv", csv_content, "text/csv")},
        )
        assert resp.status_code == 422
        data = resp.json()
        assert data["error"]["code"] == "UNSUPPORTED_DOMAIN"

    def test_invalid_file_type(self, client):
        resp = client.post(
            "/api/v1/profile-runs",
            data={"primary_domain": "Payments"},
            files={"file": ("data.json", b'{"a":1}', "application/json")},
        )
        assert resp.status_code == 422

    def test_run_not_found(self, client):
        resp = client.get("/api/v1/profile-runs/nonexistent-id")
        assert resp.status_code == 404

    def test_xlsx_single_sheet(self, client):
        """Test XLSX processing with a single sheet."""
        from openpyxl import Workbook
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.append(["id", "name", "amount"])
        for i in range(20):
            ws.append([f"R{i}", f"Name{i}", str(i * 100)])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = client.post(
            "/api/v1/profile-runs",
            data={"primary_domain": "Finance"},
            files={"file": ("report.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["status"] == "completed"

    def test_multiple_xlsx_sheets_rejected(self, client):
        """XLSX with multiple non-empty sheets should be rejected."""
        from openpyxl import Workbook
        from io import BytesIO

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["col1", "col2"])
        ws1.append(["a", "b"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["col3", "col4"])
        ws2.append(["c", "d"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = client.post(
            "/api/v1/profile-runs",
            data={"primary_domain": "Payments"},
            files={"file": ("multi.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 422
        assert "MULTIPLE_XLSX_SHEETS" in resp.json()["error"]["code"]
