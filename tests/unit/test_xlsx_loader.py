"""Tests for XLSX file loading."""

import pytest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from app.core.config import Settings
from app.core.exceptions import FileValidationError, DatasetLimitError
from app.services.ingestion.xlsx_loader import XLSXLoader


@pytest.fixture
def loader() -> XLSXLoader:
    settings = Settings(
        DATABASE_URL="postgresql://x:x@localhost/test",
        MAX_DATASET_ROWS=1000,
        MAX_DATASET_COLUMNS=50,
    )
    return XLSXLoader(settings)


@pytest.fixture
def single_sheet_xlsx(tmp_path: Path) -> Path:
    """Create a valid single-sheet XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "name", "amount"])
    ws.append([1, "Alice", 100.50])
    ws.append([2, "Bob", 200.75])
    ws.append([3, "Charlie", 50.00])
    path = tmp_path / "single.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    """Create an XLSX file with multiple non-empty sheets."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["col1", "col2"])
    ws1.append(["a", "b"])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["col3", "col4"])
    ws2.append(["c", "d"])

    path = tmp_path / "multi.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def empty_xlsx(tmp_path: Path) -> Path:
    """Create an empty XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    return path


class TestXLSXLoader:
    """Test XLSX loading and validation."""

    def test_load_single_sheet(self, loader: XLSXLoader, single_sheet_xlsx: Path):
        df = loader.load(single_sheet_xlsx)
        assert len(df) == 3
        assert list(df.columns) == ["id", "name", "amount"]

    def test_reject_multiple_sheets(self, loader: XLSXLoader, multi_sheet_xlsx: Path):
        with pytest.raises(FileValidationError) as exc_info:
            loader.load(multi_sheet_xlsx)
        assert exc_info.value.code == "MULTIPLE_XLSX_SHEETS"
        assert "Sheet1" in exc_info.value.details["sheet_names"]
        assert "Sheet2" in exc_info.value.details["sheet_names"]

    def test_reject_empty_workbook(self, loader: XLSXLoader, empty_xlsx: Path):
        with pytest.raises(FileValidationError) as exc_info:
            loader.load(empty_xlsx)
        assert exc_info.value.code == "EMPTY_WORKBOOK"

    def test_file_not_found(self, loader: XLSXLoader, tmp_path: Path):
        path = tmp_path / "nonexistent.xlsx"
        with pytest.raises(FileValidationError) as exc_info:
            loader.load(path)
        assert exc_info.value.code == "FILE_NOT_FOUND"

    def test_row_limit_exceeded(self, tmp_path: Path):
        settings = Settings(
            DATABASE_URL="postgresql://x:x@localhost/test",
            MAX_DATASET_ROWS=5,
            MAX_DATASET_COLUMNS=50,
        )
        loader = XLSXLoader(settings)

        wb = Workbook()
        ws = wb.active
        ws.append(["id", "val"])
        for i in range(10):
            ws.append([i, i * 10])
        path = tmp_path / "big.xlsx"
        wb.save(path)

        with pytest.raises(DatasetLimitError) as exc_info:
            loader.load(path)
        assert exc_info.value.code == "DATASET_ROW_LIMIT_EXCEEDED"

    def test_single_empty_sheet_plus_data_sheet(self, loader: XLSXLoader, tmp_path: Path):
        """Workbook with one empty sheet and one data sheet should load fine."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Empty"

        ws2 = wb.create_sheet("Data")
        ws2.append(["col1", "col2"])
        ws2.append(["a", "b"])
        ws2.append(["c", "d"])

        path = tmp_path / "mixed.xlsx"
        wb.save(path)

        df = loader.load(path)
        assert len(df) == 2
        assert list(df.columns) == ["col1", "col2"]

    def test_all_loaded_as_string_dtype(self, loader: XLSXLoader, single_sheet_xlsx: Path):
        df = loader.load(single_sheet_xlsx)
        for col in df.columns:
            assert df[col].dtype == object
